from openpyxl import Workbook, load_workbook
import os
from datetime import date


def export_xlsx_data(data, service_name, account_name, region_name, config, aditional_names=[]):

    today = date.today()
    date_ = today.strftime("%Y%m%d")

    if not os.path.isdir(config['result_dir']):
        os.mkdir(config['result_dir'])
    
    filename = f'{service_name}_{account_name}_{region_name}_{"_".join(aditional_names)}_{date_}.xlsx'
    filename = f'{config["result_dir"]}/{filename}'

    header_counter = 0
    column = 1
    row = 1

    if os.path.isfile(filename):
        wb = load_workbook(filename=filename)

        if region_name in wb.sheetnames:
            worksheet = wb[region_name]
        else:
            worksheet = wb.create_sheet(region_name)
        row = len(tuple(worksheet.rows)) + 2
        header_counter = 1000
    else:
        wb = Workbook()
        worksheet = wb.create_sheet(region_name)

    for i in data:

        for k, v in i.items(): 
            if header_counter < len(i.keys()):
                cell_data = worksheet.cell(row=row, column=column, value=k)
                header_counter = header_counter + 1  
            cell_data = worksheet.cell(row=row+1, column=column, value=v)       

            column = column + 1

        row = row + 1   
        column = 1

    wb.save(filename)   

    return row


